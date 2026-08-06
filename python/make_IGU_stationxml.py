#!/usr/bin/env python3
"""
Build a StationXML file for SmartSolo IGU stations from SmartSolo_recloc.xlsx.

Network code : IG
Station types: flag=0 (IGU), flag=1 (IGU_EB)
Channels     : GPZ (az=0, dip=-90), GPN (az=0, dip=0), GPE (az=90, dip=0)
Location code: 00
Sample rate  : 1000 sps
Response     : IGU_sensor-datalogger_generic_8-5-26.xml (applied to all channels)

Output written one directory above the instruments/ folder.

Validation fixes (stationxml-validator codes):
  210/310 - Reject non-datetime or pre-2024 start values (Excel serial-0
            artefacts and float campaign codes like 202411.0).
  211     - Same; eliminates bogus 1970-epoch channel overlaps.
  212     - Station endDate set to None whenever any channel is open-ended,
            so null-end channels are always consistent with the station epoch.
  222     - Reject epochs where lat==0 AND lon==0 (unfilled spreadsheet cells).
  223     - Also reject epochs where elevation==0 (same root cause).
  304     - Populate Channel.sensor with an Equipment object.
"""

import copy
import json
import re
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime as pydt

import openpyxl
import obspy
from obspy import UTCDateTime
from obspy.core.inventory import Channel, Equipment, Inventory, Network, Site, Station

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Earliest plausible deployment date.  Rejects Excel serial-0 artefacts
# (->1960-01-01) and float campaign codes (e.g. 202411.0 -> 1970-01-03).
_MIN_DATE = pydt(2024, 1, 1)

# End dates on/after this year are a spreadsheet placeholder meaning
# "still active / unknown".  Treat them as open-ended (None).
_PLACEHOLDER_END_YEAR = 2029

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = '/media/chopp/HDD1/chet-meq/cape_modern'
XLSX = f'{BASE}/instruments/SmartSolo_recloc.xlsx'
CONV_XLSX = f'{BASE}/instruments/IGU data conversion.xlsx'
RESP_XML = f'{BASE}/instruments/response/IGU_sensor-datalogger_generic_8-5-26.xml'
OUTPUT = f'{BASE}/IG_SmartSolo_IGU.xml'

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
NET_CODE = 'IG'
IGU_FLAGS = {0, 1}   # 0 = IGU, 1 = IGU_EB
LOC_CODE = ''
SAMPLE_RATE = 1000.0

# (channel_code, azimuth_deg, dip_deg)
# Dip convention: from horizontal, positive downward (FDSN/SEED)
CHANNELS = [
    ('GPZ',  0.0, -90.0),   # vertical (sensor positive up)
    ('GPN',  0.0,   0.0),   # horizontal north
    ('GPE', 90.0,   0.0),   # horizontal east
]

# Epoch column layout (0-indexed).
# Each tuple: (start_col, end_col, lon_col, lat_col, elev_col)
# Epoch 1 uses the *measured* GPS columns (8-10), not the plan columns (5-7).
# Epochs 4-15 (cols 27+) follow a regular 7-column stride.
EPOCH_COLS = [
    ( 3,  4,  8,  9, 10),       # epoch  1  (GPS-measured)
    (13, 14, 15, 16, 17),       # epoch  2
    (20, 21, 22, 23, 24),       # epoch  3
    (27, 28, 29, 30, 31),       # epoch  4
    (34, 35, 36, 37, 38),       # epoch  5
    (41, 42, 43, 44, 45),       # epoch  6
    (48, 49, 50, 51, 52),       # epoch  7
    (55, 56, 57, 58, 59),       # epoch  8
    (62, 63, 64, 65, 66),       # epoch  9
    (69, 70, 71, 72, 73),       # epoch 10
    (76, 77, 78, 79, 80),       # epoch 11
    (83, 84, 85, 86, 87),       # epoch 12
    (90, 91, 92, 93, 94),       # epoch 13
    (97, 98, 99, 100, 101),     # epoch 14
    (104, 105, 106, 107, 108),  # epoch 15
]

# Sensor equipment description (fixes validator code 304).
SENSOR_EQUIP = Equipment(
    type='Short-Period Seismograph',
    description='SmartSolo IGU-16HR 3C geophone with integrated digitizer',
    manufacturer='SmartSolo / DTCC',
    model='IGU-16HR3C',
)

# ---------------------------------------------------------------------------
# Load response template from the generic XML
# ---------------------------------------------------------------------------
resp_inv = obspy.read_inventory(RESP_XML)
_resp_template = resp_inv[0][0][0].response


def _make_response():
    """Return an independent deep copy of the response template."""
    return copy.deepcopy(_resp_template)


# ---------------------------------------------------------------------------
# Campaign start-cell parsing (blue / orange colour coding)
# ---------------------------------------------------------------------------
# Epochs 4-15 encode the campaign and sensor colour in the start cell as a
# string like '202501_blue' or '202503_orange'.  Earlier epochs (1-3) use
# plain datetime objects; those default to blue (the first-deployed sensor).
# Entries labelled '_IGU_EB' represent a different instrument and are skipped.
_CAMPAIGN_RE = re.compile(r'^(\d{4})(\d{2})_(blue|orange|IGU_EB)$')


def _load_campaign_times() -> 'dict[str, tuple[pydt | None, pydt | None]]':
    """Load Cape campaign start/end times from 'IGU data conversion.xlsx'.

    Reads the 'Cape' sheet.  Returns a dict keyed by normalised campaign
    label (e.g. '202506_blue', '202506_orange'); values are (start, end)
    pydt pairs, either of which may be None when not yet recorded.
    """
    _key_re = re.compile(r'^\d{6}(?:_Cape)?_(blue|orange|IGU_EB)$')
    wb = openpyxl.load_workbook(CONV_XLSX)
    ws = wb['Cape']
    result: dict = {}
    for row in ws.iter_rows(values_only=True):
        nodes = row[1]
        if not isinstance(nodes, str) or not _key_re.match(nodes):
            continue
        key = nodes.replace('_Cape_', '_')   # '202506_Cape_blue' → '202506_blue'
        start = row[2] if isinstance(row[2], pydt) else None
        end   = row[3] if isinstance(row[3], pydt) else None
        result[key] = (start, end)
    return result


_CAMPAIGN_TIMES = _load_campaign_times()


def _parse_end(raw):
    """Convert a spreadsheet end-date cell to UTCDateTime or None.

    Returns None for:
      - non-datetime values (float campaign codes, None)
      - placeholder end dates (year >= _PLACEHOLDER_END_YEAR)
    """
    if not isinstance(raw, pydt):
        return None
    if raw.year >= _PLACEHOLDER_END_YEAR:
        return None
    return UTCDateTime(raw)


# ---------------------------------------------------------------------------
# USGS 3DEP elevation lookup (Point Query Service)
# ---------------------------------------------------------------------------
_EPQS_URL = (
    'https://epqs.nationalmap.gov/v1/json'
    '?x={lon}&y={lat}&wkid=4326&units=Meters&includeDate=false'
)
_DEM_NODATA = -1_000_000   # sentinel returned by EPQS when no data

# Disk cache so re-runs skip the network entirely.
_DEM_CACHE_PATH = BASE + '/instruments/dem_elevation_cache.json'

def _load_dem_cache() -> dict:
    try:
        with open(_DEM_CACHE_PATH) as f:
            return {tuple(map(float, k.split(','))): v
                    for k, v in json.load(f).items()}
    except (FileNotFoundError, ValueError, KeyError):
        return {}

def _save_dem_cache(cache: dict) -> None:
    with open(_DEM_CACHE_PATH, 'w') as f:
        json.dump({f'{lat},{lon}': v for (lat, lon), v in cache.items()}, f)

_dem_cache: dict = _load_dem_cache()


def _fetch_dem(lat: float, lon: float) -> 'float | None':
    """Query the USGS 3DEP EPQS for a single point.  Returns metres or None."""
    url = _EPQS_URL.format(lat=lat, lon=lon)
    for attempt in range(3):
        try:
            with urllib.request.urlopen(url, timeout=20) as resp:
                data = json.loads(resp.read().decode())
            val = float(data['value'])
            return None if val <= _DEM_NODATA else val
        except (urllib.error.URLError, OSError, KeyError, ValueError):
            if attempt < 2:
                time.sleep(2.0)
    return None


def _prefetch_dem(coords: 'list[tuple[float, float]]') -> None:
    """Populate _dem_cache for a list of (lat, lon) pairs in parallel."""
    unique = [c for c in set(coords) if c not in _dem_cache]
    if not unique:
        print(f'DEM cache: all {len(set(coords))} elevations already cached.')
        return
    print(f'Querying USGS 3DEP for {len(unique)} missing elevations...', flush=True)
    done = 0
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(_fetch_dem, lat, lon): (lat, lon)
                   for lat, lon in unique}
        for fut in as_completed(futures):
            lat, lon = futures[fut]
            _dem_cache[(lat, lon)] = fut.result()
            done += 1
            print(f'  {done}/{len(unique)}', end='\r', flush=True)
    print(flush=True)
    _save_dem_cache(_dem_cache)
    n_ok  = sum(1 for v in _dem_cache.values() if v is not None)
    n_bad = sum(1 for v in _dem_cache.values() if v is None)
    print(f'  DEM lookup: {n_ok} elevations retrieved, {n_bad} returned no data.')


# ---------------------------------------------------------------------------
# Load spreadsheet
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(XLSX)
ws = wb['All sensors']
rows = list(ws.iter_rows(values_only=True))
data_rows = rows[1:]   # skip header

# ---------------------------------------------------------------------------
# Pre-fetch DEM elevations for stations where plan elevation is missing
# ---------------------------------------------------------------------------
_needs_dem: list[tuple[float, float]] = []
for _row in data_rows:
    if not any(v is not None for v in _row):
        continue
    _sid = _row[0]; _flag = _row[2]
    if _sid is None or _flag is None or int(_flag) not in IGU_FLAGS:
        continue
    # Primary coords: plan (cols 5-6) with GPS epoch-1 fallback (cols 8-9).
    _lon = _row[5] or _row[8]; _lat = _row[6] or _row[9]
    _elev = _row[7] if (_row[7] is not None and _row[7] != 0.0) else None
    if _lon and _lat and not (_lon == 0.0 and _lat == 0.0) and _elev is None:
        _needs_dem.append((float(_lat), float(_lon)))

_prefetch_dem(_needs_dem)

# ---------------------------------------------------------------------------
# Build Station objects
# ---------------------------------------------------------------------------
stations = []
skipped_non_igu = 0
skipped_no_epochs = 0

for row in data_rows:
    if not any(v is not None for v in row):
        continue

    station_id = row[0]
    flag = row[2]

    if station_id is None or flag is None:
        continue
    if int(flag) not in IGU_FLAGS:
        skipped_non_igu += 1
        continue

    sta_code = f'{int(station_id):04d}'

    # ------------------------------------------------------------------
    # Determine overall deployment span from all known epochs
    # ------------------------------------------------------------------
    all_starts = []
    all_ends   = []

    for (sc, ec, lonc, latc, elevc) in EPOCH_COLS:
        start_raw = row[sc]
        if isinstance(start_raw, pydt):
            if start_raw < _MIN_DATE:
                continue
            all_starts.append(UTCDateTime(start_raw))
            all_ends.append(_parse_end(row[ec]))
        elif isinstance(start_raw, str):
            m = _CAMPAIGN_RE.match(start_raw.strip())
            if not m or m.group(3) == 'IGU_EB':
                continue
            times = _CAMPAIGN_TIMES.get(start_raw.strip())
            if times is None or times[0] is None or times[0] < _MIN_DATE:
                continue
            all_starts.append(UTCDateTime(times[0]))
            all_ends.append(UTCDateTime(times[1]) if times[1] is not None else None)

    if not all_starts:
        skipped_no_epochs += 1
        continue

    sta_start = min(all_starts)
    sta_end   = None if any(e is None for e in all_ends) else max(all_ends)

    # ------------------------------------------------------------------
    # Station coordinates — prefer plan (cols 5-7), GPS epoch-1 fallback
    # ------------------------------------------------------------------
    lon  = row[5]
    lat  = row[6]
    elev = row[7]

    if lon is None or lat is None or (lon == 0.0 and lat == 0.0):
        lon  = row[8]
        lat  = row[9]
        elev = row[10]

    if lon is None or lat is None or (lon == 0.0 and lat == 0.0):
        skipped_no_epochs += 1
        continue

    if elev is None or elev == 0.0:
        elev = _dem_cache.get((float(lat), float(lon)), 0.0) or 0.0

    # ------------------------------------------------------------------
    # Single GPZ/GPN/GPE channel set spanning the full deployment
    # ------------------------------------------------------------------
    channels = []
    for (ch_code, az, dip) in CHANNELS:
        ch = Channel(
            code=ch_code,
            location_code=LOC_CODE,
            latitude=float(lat),
            longitude=float(lon),
            elevation=float(elev),
            depth=0.0,
            azimuth=az,
            dip=dip,
            sample_rate=SAMPLE_RATE,
            start_date=sta_start,
            end_date=sta_end,
        )
        ch.response = _make_response()
        ch.sensor   = copy.deepcopy(SENSOR_EQUIP)  # fixes 304
        channels.append(ch)

    sta = Station(
        code=sta_code,
        latitude=float(lat),
        longitude=float(lon),
        elevation=float(elev),
        start_date=sta_start,
        end_date=sta_end,
        site=Site(name=f'SmartSolo IGU {sta_code}'),
        channels=channels,
    )
    stations.append(sta)

# ---------------------------------------------------------------------------
# Build and write inventory
# ---------------------------------------------------------------------------
net_start = min(s.start_date for s in stations)

net = Network(
    code=NET_CODE,
    stations=stations,
    description='SmartSolo IGU nodal array (Cape / Newberry)',
    start_date=net_start,
)

inv = Inventory(
    networks=[net],
    source='SmartSolo_recloc.xlsx',
)

inv.write(OUTPUT, format='STATIONXML')

total_channels = sum(len(s.channels) for s in stations)
print(f'Written {len(stations)} stations ({total_channels} total channels) to:')
print(f'  {OUTPUT}')
print(f'Skipped: {skipped_non_igu} non-IGU rows, {skipped_no_epochs} IGU rows with no valid epochs.')
